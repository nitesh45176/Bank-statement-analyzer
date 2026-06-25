from openpyxl import Workbook
from openpyxl.styles import Font

class ExcelExporter:

    @staticmethod
    def export(
        account_details,
        transactions,
        summary,
        monthly_summary,
        category_summary,
        output_path: str
    ):

        workbook = Workbook()

        # ==================================================
        # Sheet 1 : Account Details
        # ==================================================

        account_sheet = workbook.active
        account_sheet.title = "Account Details"

        account_sheet["A1"] = "Field"
        account_sheet["B1"] = "Value"

        account_sheet["A1"].font = Font(bold=True)
        account_sheet["B1"].font = Font(bold=True)

        row = 2

        for key, value in account_details.items():

            account_sheet.cell(row=row, column=1).value = key.replace("_", " ").title()
            account_sheet.cell(row=row, column=2).value = value

            row += 1


        account_sheet.append([])

        account_sheet.append([
            "Opening Balance",
            account_details.get("opening_balance", "")
        ])

        account_sheet.append([
            "Closing Balance",
            summary["closing_balance"]
        ])

        account_sheet.append([
            "Credit Count",
            summary["credit_count"]
        ])

        account_sheet.append([
            "Credit Amount",
            summary["total_credit"]
        ])

        account_sheet.append([
            "Debit Count",
            summary["debit_count"]
        ])

        account_sheet.append([
            "Debit Amount",
            summary["total_debit"]
        ])

        account_sheet.append([
            "Categorized %",
            summary["categorized_percent"]
        ])

        account_sheet.append([
            "Uncategorized %",
            summary["uncategorized_percent"]
        ])
        account_sheet.append(["Total Credit", summary["total_credit"]])
        account_sheet.append(["Total Debit", summary["total_debit"]])
        account_sheet.append(["Closing Balance", summary["closing_balance"]])

        # ==================================================
        # Sheet 2 : Transaction Ledger
        # ==================================================

        ledger = workbook.create_sheet("Transaction Ledger")

        headers = [
            "Date",
            "Value Date",
            "Narration",
            "Reference",
            "Debit",
            "Credit",
            "Balance",
            "Category"
        ]

        ledger.append(headers)

        for cell in ledger[1]:
            cell.font = Font(bold=True)

        for transaction in transactions:

            ledger.append([
                transaction.date,
                transaction.value_date,
                transaction.narration,
                transaction.reference,
                transaction.debit,
                transaction.credit,
                transaction.balance,
                transaction.category
            ])

        # ==================================================
        # Sheet 3 : Analytics
        # ==================================================

        analytics = workbook.create_sheet("Analytics")

        analytics["A1"] = "Month"
        analytics["B1"] = "Credit"
        analytics["C1"] = "Debit"
        analytics["D1"] = "Transactions"

        for cell in analytics[1]:
            cell.font = Font(bold=True)

        row = 2

        for month, values in monthly_summary.items():

            analytics.cell(row=row, column=1).value = month
            analytics.cell(row=row, column=2).value = values["credit"]
            analytics.cell(row=row, column=3).value = values["debit"]
            analytics.cell(row=row, column=4).value = values["count"]

            row += 1

        row += 2

        analytics.cell(row=row, column=1).value = "Category Summary"
        analytics.cell(row=row, column=1).font = Font(bold=True)

        row += 1

        analytics.cell(row=row, column=1).value = "Category"
        analytics.cell(row=row, column=2).value = "Debit"
        analytics.cell(row=row, column=3).value = "Credit"
        analytics.cell(row=row, column=4).value = "Count"

        for cell in analytics[row]:
            cell.font = Font(bold=True)

        row += 1

        for category, values in category_summary.items():

            analytics.cell(row=row, column=1).value = category
            analytics.cell(row=row, column=2).value = values["debit"]
            analytics.cell(row=row, column=3).value = values["credit"]
            analytics.cell(row=row, column=4).value = values["count"]

            row += 1


        # Format every worksheet
        for sheet in workbook.worksheets:

            # Freeze header row
            sheet.freeze_panes = "A2"

            # Auto column width
            for column in sheet.columns:

                max_length = 0

                column_letter = column[0].column_letter

                for cell in column:

                    if cell.value is not None:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                    # Format numbers
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

                sheet.column_dimensions[column_letter].width = min(
                    max_length + 3,
                    60
                )
        workbook.save(output_path)